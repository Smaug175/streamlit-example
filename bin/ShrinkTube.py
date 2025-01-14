import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from bin.model.Normal.CuttingClipMold import DC0124_AD03, DC0124_AD03_, DC0121_AD03, DC0121_AD04, DC0125_AD07
from bin.model.Normal.FormingMandrel import DC0121_SS01, DC0124_SS01, DC0125_SS01
from bin.model.Normal.FormingMold import DIEO
from bin.model.Normal.FormingUnloadingMold import DC0124_AD02, DC0121_AD02, DC0125_AD06_F
from bin.model.Normal.ShrinkTubeMandrel import ADBT
from bin.model.Normal.ShrinkTubeMold import ADIE
from bin.model.Normal.ShrinkTubeUnloadingMold import DC0124_AD01, DC0121_AD01, DC0125_AD06_S
from bin.model.TubeModel import TubeModelClass

import shutil
import zipfile
from bin.utils.SQLite_control import MoldControl

class ShrinkTubeClass:
    """缩管计算类

    这个类实现普通抽抽管的所有功能，是一个完整的抽管计算类。

    Methods
    _______
    get_tube_params() -> dict:
        将管件参数转换为字典，并返回
    __init__(logger, file_path: str):
        初始化ShrinkTubeClass
    get_numbers(logger, mold_name, machine_type) -> str:
        通过查询数据库，获取图号
    calculate(user_name: str, config_setting_instance, Normal_Add: bool, mold_list: list, machine_type: str) -> dict:
        计算模具的参数
    modify_parameters(mold_name: str, key: str, value: str):
        修改模具的参数
    df_to_dict(df: pd.DataFrame) -> dict:
        将DataFrame转换为字典
    get_all_params() -> dict:
        获取当前计算的所有参数，返回嵌套字典
    save_all():
        将所有参数保存到数据库中
    output_dxf(output_path: str):
        将所有模具的DXF文件保存到指定路径
    output_excel(output_path: str):
        将所有模具的参数保存到Excel文件中
    """
    def _get_params_from_tube(self, params: pd.DataFrame) -> tuple:
        """从管件参数中提取参数"""

        params = params.to_dict(orient='list')
        params = dict(zip(params['Parameter'], params['Value']))
        D = params['D']
        Tx = {}
        for item in params:
            if item.startswith('T'):
                Tx[item] = params[item]
        L = params['L']
        Mx = {}
        for item in params:
            if item.startswith('M'):
                Mx[item] = params[item]
        Lx = {}
        for item in params:
            if item.startswith('L') and item != 'L' and item != 'D':
                Lx[item] = params[item]

        self.logger.info('管件参数提取成功，位置位于：ShrinkTubeClass._get_params_from_tube')
        return L, D, Lx, Mx, Tx

    def get_tube_params(self) -> dict:
        """将管件参数转换为字典，并返回"""
        return self.df_to_dict(self.tube_df_params)

    def get_tube_params_df(self):
        """增加参数描述"""
        parameter_list = list(self.tube_df_params['Parameter'])
        value_list = []
        for v in self.tube_df_params['Value']:
            value_list.append(str(v))
        description_list = []
        for param in parameter_list:
            description_list.append(self.Parameter_Description['管件参数'][param])
        tube_params_df = pd.DataFrame({
            '参数': parameter_list,
            '值': value_list,
            '描述': description_list
        })
        return tube_params_df

    def __init__(self, logger, file_path):
        """初始化ShrinkTubeClass"""
        self.logger = logger
        # 导入模具的参数描述和计算方法
        with open('bin/model/Normal/Parameter_Description.json', 'r', encoding='utf-8') as file:
            self.Parameter_Description = json.load(file)
        with open('bin/model/Normal/Parameter_Calculate_Method.json', 'r', encoding='utf-8') as file:
            self.Parameter_Calculate_Method = json.load(file)

        if file_path:
            self.load_tube(file_path)
        else:
            print('未传入管件参数，无法计算，请使用load_tube()函数加载文件')

    def load_tube(self, file_path: str):
        self.file_name = file_path.split('/')[-1][:-4]
        tube = TubeModelClass(file_path, self.logger)
        self.tube_df_params = tube.get_params()
        # 额外信息
        self.external_params = {
            '图号': '/',
            '件数': '/',
            '车种规格': self.get_tube_params()['车种规格'],  # 从输入图中读取该参数
            '设计者': '/',  # 来自操作的用户登录的信息
        }

    def get_numbers(self, logger, mold_name, machine_type):
        """通过查询数据库，获取图号"""
        # TODO: 不同的机器的图号不同，需要根据机器的名称来获取

        if machine_type == 'DC0124':
            self.Drawing_Number = {
                '裁剪夹模': 'AD03',
                '成型芯轴': 'SS01',
                '成型模': 'DIEO',
                '成型退料模': 'AD02',
                '缩管模': 'ADIE',
                '抽管芯轴': 'ADBT',
                '抽管退料模': 'AD01',
            }
        elif machine_type == 'DC0121':
            self.Drawing_Number = {
                '裁剪夹模1': 'AD03',
                '裁剪夹模2': 'AD04',
                '成型芯轴': 'SS01',
                '成型模': 'DIEO',
                '成型退料模': 'AD02',
                '缩管模': 'ADIE',
                '抽管芯轴': 'ADBT',
                '抽管退料模': 'AD01',
            }
        elif machine_type == 'DC0125':
            self.Drawing_Number = {
                '裁剪夹模': 'AD07',
                '成型芯轴': 'SS01',
                '成型模': 'DIEO',
                '成型退料模': 'AD06_F',
                '缩管模': 'ADIE',
                '抽管芯轴': 'ADBT',
                '抽管退料模': 'AD06_S',
            }

        # 获取新的图号
        # crypto_data = CryptoDataClass(logger)

        try:
            # machine_type
            # big_graph_number
            sqlite_control = MoldControl(self.logger)
            number = sqlite_control.query_max_graph_number(machine_type, self.Drawing_Number[mold_name])
            # 如果不存在记录，则会返回 0
            
            # name = machine_type+'-'+self.Drawing_Number[mold_name]
            # print(name)
            # data_dict = crypto_data.decrypt_file_by_big_graph_number(name)

            # # print('获取{}的图号，数据字典为：'.format(mold_name))
            # # pprint(data_dict.keys())

            # keys = data_dict.keys()
            # index = []
            # for key in keys:
            #     index.append(int(key[-4:]))
                
            number = number + 1
            str_number = str(number)

            if len(str_number) != 4:
                new_number = self.Drawing_Number[mold_name] + '0'*(4-len(str_number)) + str_number
            else:
                new_number = self.Drawing_Number[mold_name] + str_number
            return machine_type + '-' + new_number
        except:
            print('没有找到{}的图号，将使用默认的图号'.format(mold_name))
            self.logger.error('没有找到{}的图号，将使用默认的图号'.format(mold_name))
            new_number = self.Drawing_Number[mold_name] + '0000'
            return machine_type + '-' + new_number

    def calculate(self, user_name: str, config_setting_instance, Normal_Add: bool, mold_list: list, machine_type: str) -> dict:
        """计算模具的参数"""

        self.Mold_Object = {}

        if machine_type == 'DC0124':
            # 不同的机器类型，其件数不同
            # 在这个机器类型中，Normal_Add为True则裁剪夹模只要4个，False则裁剪夹模需要不同尺寸的分别2个
            # 裁剪夹模：正常的尺寸。
            # 裁剪夹模_：不要加0.3，多出来的尺寸。
            # 这两个东西共用同一个模具和图号，只是参数不同。
            L, D, Lx, Mx, Tx = self._get_params_from_tube(self.tube_df_params)

            if Normal_Add or len(Tx) < 3:
                ccm = 4
            else:
                ccm = 2
            self.Mold_Counts = {
                '裁剪夹模': ccm,
                '成型芯轴': 1,
                '成型模': 1,
                '成型退料模': 1,
                '缩管模': 1,
                '抽管芯轴': 1,
                '抽管退料模': 1,
            }
            self.Mold_Object = {
                        '裁剪夹模': DC0124_AD03(self.logger, config_setting_instance),
                        '成型模': DIEO(self.logger, config_setting_instance),
                        '成型芯轴': DC0124_SS01(self.logger, config_setting_instance),
                        '成型退料模': DC0124_AD02(self.logger, config_setting_instance),
                        '缩管模': ADIE(self.logger, config_setting_instance),
                        '抽管芯轴': ADBT(self.logger, config_setting_instance),
                        '抽管退料模': DC0124_AD01(self.logger, config_setting_instance),
                    }

            mold_name = ['裁剪夹模', '成型模', '成型芯轴', '成型退料模', '缩管模', '抽管芯轴', '抽管退料模']

            for i in range(len(mold_name)):
                if mold_name[i] not in mold_list:
                    self.Mold_Object.pop(mold_name[i]) # 删除不需要的模具

            # 用于记录裁剪夹模_的图号
            temp_number = ''

            # 需要输入额外参数
            for mold_name in self.Mold_Object:
                mold = self.Mold_Object[mold_name]
                # 获取新的图号，不同的模具有不同的图号，需要根据模具的名称来获取
                self.external_params['图号'] = self.get_numbers(self.logger, mold_name, machine_type)
                if mold_name == '裁剪夹模':
                    temp_number = self.external_params['图号']
                self.external_params['件数'] = self.Mold_Counts[mold_name]
                self.external_params['设计者'] = user_name
                # 设置参数，Normal_Add为True表示加0.3，False表示不加0.3
                mold.set_params(self.tube_df_params, self.external_params, Normal_Add)

            if not Normal_Add and len(Tx) > 2 and '裁剪夹模' in mold_list:
                # 计算额外的裁剪夹模
                # 裁剪夹模_的参数和裁剪夹模一样，只是不加0.3
                self.Mold_Object['裁剪夹模_'] = DC0124_AD03_(self.logger, config_setting_instance)
                # 这里的图号，是根据上面给的图号来的
                before = temp_number[:-4]
                after = temp_number[-4:]
                new_number = str(int(after) + 1)
                if len(new_number) != 4:
                    new_number = before + '0'*(4-len(new_number)) + new_number
                else:
                    new_number = before + str(int(after) + 1)
                self.external_params['图号'] = new_number
                self.external_params['件数'] = 2
                self.external_params['设计者'] = user_name
                self.Mold_Object['裁剪夹模_'].set_params(self.tube_df_params, self.external_params, Normal_Add)

            self.logger.info('抽管计算成功')
        elif machine_type == 'DC0121':
            L, D, Lx, Mx, Tx = self._get_params_from_tube(self.tube_df_params)

            self.Mold_Counts = {
                '裁剪夹模1': 2,
                '裁剪夹模2': 2,
                '成型芯轴': 1,
                '成型模': 1,
                '成型退料模': 1,
                '缩管模': 1,
                '抽管芯轴': 1,
                '抽管退料模': 1,
            }
            self.Mold_Object = {
                '裁剪夹模1': DC0121_AD03(self.logger, config_setting_instance),
                '裁剪夹模2': DC0121_AD04(self.logger, config_setting_instance),
                '成型模': DIEO(self.logger, config_setting_instance),
                '成型芯轴': DC0121_SS01(self.logger, config_setting_instance),
                '成型退料模': DC0121_AD02(self.logger, config_setting_instance),
                '缩管模': ADIE(self.logger, config_setting_instance),
                '抽管芯轴': ADBT(self.logger, config_setting_instance),
                '抽管退料模': DC0121_AD01(self.logger, config_setting_instance),
            }

            mold_name = ['裁剪夹模1', '裁剪夹模2', '成型模', '成型芯轴', '成型退料模', '缩管模', '抽管芯轴', '抽管退料模']

            for i in range(len(mold_name)):
                if mold_name[i] not in mold_list:
                    self.Mold_Object.pop(mold_name[i])  # 删除不需要的模具

            # 需要输入额外参数
            for mold_name in self.Mold_Object:
                mold = self.Mold_Object[mold_name]
                # 获取新的图号，不同的模具有不同的图号，需要根据模具的名称来获取
                self.external_params['图号'] = self.get_numbers(self.logger, mold_name, machine_type)
                self.external_params['件数'] = self.Mold_Counts[mold_name]
                self.external_params['设计者'] = user_name
                # 设置参数，Normal_Add为True表示加0.3，False表示不加0.3
                mold.set_params(self.tube_df_params, self.external_params, Normal_Add)
        elif machine_type == 'DC0125':
            L, D, Lx, Mx, Tx = self._get_params_from_tube(self.tube_df_params)

            self.Mold_Counts = {
                '裁剪夹模': 2,
                '成型芯轴': 2,
                '成型模': 2,
                '成型退料模': 2,
                '缩管模': 2,
                '抽管芯轴': 2,
                '抽管退料模': 2,
            }
            self.Mold_Object = {
                '裁剪夹模': DC0125_AD07(self.logger, config_setting_instance),
                '成型模': DIEO(self.logger, config_setting_instance),
                '成型芯轴': DC0125_SS01(self.logger, config_setting_instance),
                '成型退料模': DC0125_AD06_F(self.logger, config_setting_instance),
                '缩管模': ADIE(self.logger, config_setting_instance),
                '抽管芯轴': ADBT(self.logger, config_setting_instance),
                '抽管退料模': DC0125_AD06_S(self.logger, config_setting_instance),
            }

            mold_name = ['裁剪夹模', '成型模', '成型芯轴', '成型退料模', '缩管模', '抽管芯轴', '抽管退料模']

            for i in range(len(mold_name)):
                if mold_name[i] not in mold_list:
                    self.Mold_Object.pop(mold_name[i])  # 删除不需要的模具

            # 需要输入额外参数
            for mold_name in self.Mold_Object:
                mold = self.Mold_Object[mold_name]
                # 获取新的图号，不同的模具有不同的图号，需要根据模具的名称来获取
                if mold_name == '成型退料模':
                    temp_number = self.get_numbers(self.logger, mold_name, machine_type)
                    before = temp_number[:-4]
                    after = temp_number[-4:]
                    new_number = str(int(after) + 1)
                    if len(new_number) != 4:
                        new_number = before + '0' * (4 - len(new_number)) + new_number
                    else:
                        new_number = before + str(int(after) + 1)
                    self.external_params['图号'] = new_number
                else:
                    self.external_params['图号'] = self.get_numbers(self.logger, mold_name, machine_type)
                self.external_params['件数'] = self.Mold_Counts[mold_name]
                self.external_params['设计者'] = user_name
                # 设置参数，Normal_Add为True表示加0.3，False表示不加0.3
                mold.set_params(self.tube_df_params, self.external_params, Normal_Add)

        return self.Mold_Object

    def modify_parameters(self, mold_name: str, key: str, value: str):
        """修改模具的参数"""
        # print('mold_name:', mold_name)
        # print('key:', key)
        # print('value:', value)
        try:
            mold = self.Mold_Object[mold_name]
            mold._modify_parameters(key, value)
        except Exception as e:
            temp_str = 'Error: 修改参数失败，位置位于：ShrinkTubeClass.modify_parameters'
            self.logger.error(temp_str)

    def df_to_dict(self, df: pd.DataFrame) -> dict:
        """将DataFrame转换为字典"""
        Dict = {}
        for i in range(len(df)):
            Dict[df['Parameter'][i]] = df['Value'][i]
        return Dict

    def get_all_params(self) -> dict:
        """获取当前计算的所有参数，返回嵌套字典"""
        # 将关键参数同时放入该字典中
        ALL_params = {
            '管件参数': self.df_to_dict(self.tube_df_params),
        }
        for mold_name in self.Mold_Object:
            mold = self.Mold_Object[mold_name]
            mold_params = mold.get_params()
            # 使用计算之后的图号作为关键key
            graph_number = self.df_to_dict(mold_params)['图号']
            ALL_params[graph_number] = self.df_to_dict(mold_params)

        return ALL_params

    def get_molds_params_df(self):
        """获取当前计算的所有模具的参数，返回 df"""
        # 将关键参数同时放入该字典中
        ALL_params = {
        }

        for mold_name in self.Mold_Object:
            mold = self.Mold_Object[mold_name]
            mold_params = mold.get_params()
            image_number = mold_params.loc[mold_params['Parameter'] == '图号', 'Value'].values[0][:-4]
            # print(image_number)
            mold_name = self.df_to_dict(mold_params)['模具名称']

            parameter_list = list(mold_params['Parameter'])
            value_list = []
            for v in mold_params['Value']:
                value_list.append(str(v))
            description_list = []
            calclater_list = []
            for param in parameter_list:
                description_list.append(self.Parameter_Description[image_number][param])
                calclater_list.append(self.Parameter_Calculate_Method[image_number][param])

            params_df = pd.DataFrame({
                '参数': parameter_list,
                '值': value_list,
                '描述': description_list,
                '计算方法': calclater_list
            })
            ALL_params[mold_name] = params_df

        return ALL_params

    def save_all(self):
        """将所有参数保存到数据库中"""
        self.get_all_params()
        #pprint(self.get_all_params())

        # crypto_data = CryptoDataClass(self.logger)
        # crypto_data.encrypt_data(self.get_all_params())

        # 使用最新的方式保存数据
        sqlite_control = MoldControl(self.logger)
        sqlite_control.insert_data(self.get_all_params())

    def output_dxf(self, output_path: str):
        """将所有模具的DXF文件保存到指定路径"""

        for mold_name in self.Mold_Object:
            mold = self.Mold_Object[mold_name]
            mold.modify_dxf() # 修改DXF文件
            mold_params = mold.get_params()
            mold_params_dict = self.df_to_dict(mold_params)
            mold.save_dxf(output_path, mold_params_dict['图号'])
        self.logger.info('所有模具的DXF文件保存成功')

    def output_excel(self, output_path: str):
        """将所有模具的参数保存到Excel文件中"""

        #已经存在该文件，则删除
        excel_filename = os.path.join(output_path, '模具参数汇总.xlsx')
        if os.path.exists(excel_filename):
            os.remove(excel_filename)

        # 检查文件是否存在，如果不存在则创建
        if not os.path.exists(excel_filename):
            with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='w') as writer:
                self.tube_df_params.to_excel(writer, sheet_name='管件参数', index=False)
                for mold_name in self.Mold_Object:
                    mold = self.Mold_Object[mold_name]
                    mold_params = mold.get_params()
                    mold_params.to_excel(writer, sheet_name=mold_name, index=False)

        # 加载Excel文件
        wb = load_workbook(excel_filename)

        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            # 获取当前工作表
            ws = wb[sheet_name]
            # 设置列宽
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            # 遍历当前工作表的所有单元格，设置居中对齐
            for row in ws.iter_rows(min_row=1, min_col=1, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(excel_filename)
        wb.close()

        self.logger.info('所有模具的excel文件保存成功， 路径位于：{}'.format(excel_filename))

    def output_zip_from_cache(self,out_root):
        """将所有模具的DXF文件保存到指定路径"""
        if not os.path.exists(out_root):
            os.mkdir(out_root)
        output_path = os.path.join(out_root, self.get_tube_params()['车种规格'])
        zip_filename = os.path.join(out_root, self.get_tube_params()['车种规格'] + '.zip')

        if not os.path.exists(output_path):
            os.mkdir(output_path)
        else:
            shutil.rmtree(output_path)
            os.mkdir(output_path)

        if os.path.exists(zip_filename):
            os.remove(zip_filename)

        self.output_dxf(output_path)
        self.output_excel(output_path)

        with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(output_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    # 计算在压缩文件中的相对路径
                    arcname = os.path.relpath(file_path, output_path)
                    zipf.write(file_path, arcname)

        return zip_filename

